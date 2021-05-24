from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, NamedStyle
from copy import copy
from django.conf import settings
from django.http import HttpResponse
import os
import io
import base64
import json
import time
######################
#
# Loading and manipulating data on excel worksheet is trick, if we just search our desired by
# cell and column in for loop, it will consume memory drastically
# Hence, we need to specific access zone as specific as possible
# in theory, if the excel file 50MB, it will consume by 50 times (2.5GB) when using openpyxl
# ref: https://openpyxl.readthedocs.io/en/stable/performance.html
#
######################

class CellPosition:
  def __init__(self, row, col):
    self.row = row
    self.col = col

class TableRange:
  def __init__(self, start_row, start_col, end_row, end_col):
    self.start_cell = CellPosition(start_row, start_col)
    self.end_cell = CellPosition(end_row, end_col)

def copy_cell(source_cell, dest_pos, dest_sheet):
  dest_sheet.cell(dest_pos.row, dest_pos.col).value = source_cell.value
  if source_cell.has_style:
      dest_sheet.cell(dest_pos.row, dest_pos.col)._style = copy(source_cell._style)
  return dest_sheet.cell(dest_pos.row, dest_pos.col)

def copy_range(source_sheet, dest_sheet, copy_range, dest_anchor):
  start_row = copy_range.start_cell.row
  start_col = copy_range.start_cell.col
  end_row = copy_range.end_cell.row
  end_col = copy_range.end_cell.col
  row_count = 0 
  for row in range (start_row, end_row+1):
    col_count = 0
    for col in range (start_col, end_col + 1):
      copy_cell(source_sheet.cell(row, col), CellPosition(dest_anchor.row + row_count, dest_anchor.col + col_count) , dest_sheet)
      col_count += 1
    row_count += 1

def load_tables(workbook):
  instruction_ws = workbook["instruction"]
  num_of_table = instruction_ws['B1'].value
  first_table_inst_row = 2
  first_table_inst_col = 2
  table_index = 0
  table_list = []
  while table_index < num_of_table:
    table = {}
    table["name"] = instruction_ws.cell(first_table_inst_row + table_index , first_table_inst_col).value
    table["range"] = TableRange(
      instruction_ws.cell(first_table_inst_row + table_index , first_table_inst_col + 1).value,
      instruction_ws.cell(first_table_inst_row + table_index , first_table_inst_col + 2).value, 
      instruction_ws.cell(first_table_inst_row + table_index , first_table_inst_col + 3).value, 
      instruction_ws.cell(first_table_inst_row + table_index , first_table_inst_col + 4).value 
    )
    table["holder"] = CellPosition(
      instruction_ws.cell(first_table_inst_row + table_index , first_table_inst_col +5).value,
      instruction_ws.cell(first_table_inst_row + table_index , first_table_inst_col +6).value
    )
    table["max_row"] = instruction_ws.cell(first_table_inst_row + table_index , first_table_inst_col + 7).value

    table_index +=1
    table_list.append(table)
  return table_list

def loadWorkbook(name):
  wb_path =  os.path.join(settings.EXCEL_TEMPLATE, name)
  wb = load_workbook(wb_path)
  return wb

def saveWorkbook(wb):
  temp = os.path.join(settings.EXCEL_TEMPLATE, "debug", str(time.time())+ ".xlsx")
  if not os.path.exists(os.path.join(settings.EXCEL_TEMPLATE, "debug")):
    os.makedirs(os.path.join(settings.EXCEL_TEMPLATE, "debug"))
  wb.save(temp)

def createHttpRespond(workbook):
  bin_stream = base64.b64encode(save_virtual_workbook(workbook)).decode()
  respond_data = {
    "data" : bin_stream
  }
  response = HttpResponse(
    json.dumps(respond_data), 
    content_type='application/json')
  return response
  
def findTableByName(name, list_to_find):
  for t in list_to_find:
    if t["name"] == name:
      return t

def create_table(ws, display_name, range):
    tab = Table(displayName=display_name, ref=range)
    tab.headerRowCount = 1
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=False, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

def style_range(ws, cell_range, style):
    cells = ws[cell_range]
    for cell in cells:
        cell[0].style = style

def style_range_merge(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill